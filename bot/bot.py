import subprocess, os, re
from datetime import datetime, date
import openpyxl
from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, filters, ContextTypes

TELEGRAM_TOKEN = os.environ.get('TELEGRAM_TOKEN', '8852326670:AAEItInxXpi-Ynyoc0MvmwDVTZFJtw5o0WI')
ANTHROPIC_KEY = os.environ.get('ANTHROPIC_API_KEY', '')
GROUP_CHAT_ID = -1003951496246
DATA_FILE = '/opt/tgbot/colleagues.xlsx'

def parse_date(val):
    if not val:
        return None
    s = str(val).strip().replace(' ', '')
    # Формати: 01.06, 01.06.96, 01.06.1996, 01.06.2024
    patterns = [
        (r'^(\d{1,2})\.(\d{1,2})\.(\d{4})$', '%d.%m.%Y'),
        (r'^(\d{1,2})\.(\d{1,2})\.(\d{2})$', '%d.%m.%y'),
        (r'^(\d{1,2})\.(\d{1,2})$', None),  # тільки день.місяць
    ]
    for pattern, fmt in patterns:
        if re.match(pattern, s):
            if fmt:
                try:
                    return datetime.strptime(s, fmt).date()
                except:
                    pass
            else:
                # тільки день.місяць — додаємо поточний рік
                try:
                    parts = s.split('.')
                    return date(date.today().year, int(parts[1]), int(parts[0]))
                except:
                    pass
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None

def days_until(d):
    today = date.today()
    try:
        next_event = d.replace(year=today.year)
    except ValueError:
        next_event = d.replace(year=today.year, day=28)
    if next_event < today:
        next_event = next_event.replace(year=today.year + 1)
    return (next_event - today).days

def load_events():
    if not os.path.exists(DATA_FILE):
        return []
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    events = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip() if row[0] else ''
            reason = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            date_val = row[5] if len(row) > 5 else None
            d = parse_date(date_val)
            if name and reason and d:
                events.append({
                    'name': name,
                    'reason': reason,
                    'date': d,
                    'telegram': str(row[3]).strip() if len(row) > 3 and row[3] else '',
                })
    return events

def make_message(event, days_left):
    name = event['name']
    reason = event['reason']
    tg = event['telegram']
    timing = '🎉 СЬОГОДНІ!' if days_left == 0 else f'⏰ Через {days_left} днів'

    # Визначаємо емодзі за приводом
    r = reason.lower()
    if 'народження' in r and ('доньк' in r or 'сина' in r or 'дитин' in r):
        emoji = '👶'
    elif 'народження' in r:
        emoji = '🎂'
    elif 'річниця' in r or 'заснування' in r:
        emoji = '🏆'
    else:
        emoji = '🎊'

    msg = f'{emoji} {reason} — {name}\n{timing}'
    if tg:
        msg += f'\n📱 {tg}'
    return msg

async def do_check(bot, notify_days=[0, 7]):
    events = load_events()
    count = 0
    for e in events:
        d = days_until(e['date'])
        if d in notify_days:
            await bot.send_message(chat_id=GROUP_CHAT_ID, text=make_message(e, d))
            count += 1
    return count

async def handle(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return

    msg_text = update.message.text or ''
    chat_id = update.message.chat.id

    if msg_text == '/id':
        await update.message.reply_text(f'Chat ID: {chat_id}')
        return

    if msg_text.startswith('/delete '):
        name_to_delete = msg_text[8:].strip()
        if not os.path.exists(DATA_FILE):
            await update.message.reply_text('Файл не знайдено.')
            return
        wb = openpyxl.load_workbook(DATA_FILE)
        deleted = 0
        for sheet in wb.worksheets:
            rows_to_delete = []
            for row in sheet.iter_rows(min_row=2):
                if row[0].value and name_to_delete.lower() in str(row[0].value).lower():
                    rows_to_delete.append(row[0].row)
            for row_num in reversed(rows_to_delete):
                sheet.delete_rows(row_num)
                deleted += 1
        wb.save(DATA_FILE)
        if deleted:
            await update.message.reply_text(f'✅ Видалено {deleted} запис(ів) для "{name_to_delete}".')
        else:
            await update.message.reply_text(f'❌ Не знайдено "{name_to_delete}" у файлі.')
        return

    if msg_text.startswith('/add'):
        text = msg_text[4:].strip()
        if not text:
            await update.message.reply_text(
                'Надішли інформацію після /add:\n\n'
                '/add Іванова Марія\n'
                'ДН: 18.05.1986\n'
                'Річниця: 01.06.2024\n'
                'Посада: Агент\n'
                'Офіс: Київ\n'
                'Хобі: йога\n'
                'Квіти: троянди\n'
                'Торт: медовик\n'
                '@username'
            )
            return

        lines = text.split('\n')

        # Знаходимо ім'я — перший рядок або рядок без ключових слів
        name = ''
        birthday = None
        work_ann = None
        position = ''
        office = ''
        hobbies = ''
        flowers = ''
        cake = ''
        telegram = ''
        children = ''

        for line in lines:
            l = line.strip()
            if not l:
                continue
            ll = l.lower()

            # Дата народження
            if any(k in ll for k in ['дн:', 'народж', 'birthday', 'дата народження', 'д.н']):
                dates = re.findall(r'\d{1,2}[\.\-\s]+\d{1,2}[\.\-\s]+\d{2,4}', l)
                if dates:
                    d_str = re.sub(r'[\s\-]+', '.', dates[0])
                    birthday = parse_date(d_str)

            # Річниця
            elif any(k in ll for k in ['річниця', 'anniversary', 'прийнят']):
                dates = re.findall(r'\d{1,2}[\.\-\s]+\d{1,2}[\.\-\s]+\d{2,4}', l)
                if dates:
                    d_str = re.sub(r'[\s\-]+', '.', dates[0])
                    work_ann = parse_date(d_str)

            # Посада
            elif any(k in ll for k in ['посада', 'position', 'роль']):
                position = re.sub(r'.*?:\s*', '', l, count=1).strip()

            # Офіс
            elif any(k in ll for k in ['офіс', 'office', 'місто']):
                office = re.sub(r'.*?:\s*', '', l, count=1).strip()

            # Хобі
            elif any(k in ll for k in ['хобі', 'hobby', 'захоплен', 'інтерес']):
                hobbies = re.sub(r'.*?:\s*', '', l, count=1).strip()

            # Квіти
            elif any(k in ll for k in ['квіт', 'flower']):
                flowers = re.sub(r'.*?:\s*', '', l, count=1).strip()

            # Торт
            elif any(k in ll for k in ['торт', 'cake', 'десерт']):
                cake = re.sub(r'.*?:\s*', '', l, count=1).strip()

            # Діти
            elif any(k in ll for k in ['діт', 'child', 'син', 'дочк', 'доньк']):
                children = l

            # Телеграм
            elif '@' in l:
                telegram = (telegram + ', ' + l).strip(', ')

            # Ім'я — рядок з великої літери без спецсимволів і без дат
            elif not name and re.match(r'^[А-ЯІЇЄA-Z][а-яіїєa-zA-Z\s\'\-]+$', l) and len(l.split()) >= 2:
                name = l

        if not name:
            name = lines[0].strip() if lines else 'Невідомо'

        if not os.path.exists(DATA_FILE):
            await update.message.reply_text('❌ Файл не знайдено. Спочатку завантаж Excel.')
            return

        wb = openpyxl.load_workbook(DATA_FILE)
        ws = wb.worksheets[0]
        first_name = name.split()[1] if len(name.split()) > 1 else name

        rows_added = 0
        if birthday:
            ws.append([name, office, position, telegram,
                       f'день народження {first_name}', birthday])
            rows_added += 1
        if work_ann:
            ws.append([name, office, position, telegram,
                       'річниця роботи в ЕІ', work_ann])
            rows_added += 1
        if children and children.lower() not in ['немає', 'нема', 'no', '-']:
            ws.append([name, office, position, telegram,
                       f'день народження дітей', children])
            rows_added += 1

        wb.save(DATA_FILE)

        summary = f'✅ Додано {name}!\n\n'
        summary += f'📅 День народження: {birthday.strftime("%d.%m.%Y") if birthday else "не вказано"}\n'
        summary += f'🏆 Річниця роботи: {work_ann.strftime("%d.%m.%Y") if work_ann else "не вказано"}\n'
        summary += f'🌸 Квіти: {flowers or "не вказано"}\n'
        summary += f'🍰 Торт: {cake or "не вказано"}\n'
        summary += f'🎯 Хобі: {hobbies or "не вказано"}\n'
        summary += f'📱 Телеграм: {telegram or "не вказано"}'
        await update.message.reply_text(summary)
        return

    if msg_text == '/check':
        await update.message.reply_text('Перевіряю календар...')
        count = await do_check(ctx.bot)
        await update.message.reply_text(f'Готово! Надіслано {count} нагадувань.')
        return

    if msg_text == '/today':
        await update.message.reply_text('Події на сьогодні і через 7 днів:')
        events = load_events()
        found = []
        for e in events:
            d = days_until(e['date'])
            if d in [0, 7]:
                found.append(make_message(e, d))
        if found:
            for msg in found:
                await update.message.reply_text(msg)
        else:
            await update.message.reply_text('Немає подій на найближчі 7 днів.')
        return

    if msg_text == '/list':
        events = load_events()
        if not events:
            await update.message.reply_text('Список порожній. Завантаж Excel файл.')
            return
        lines = [f'📋 Всього подій: {len(events)}\n']
        for e in events:
            lines.append(f"• {e['name']} — {e['reason']} ({e['date'].strftime('%d.%m')})")
        text = '\n'.join(lines)
        for i in range(0, len(text), 4000):
            await update.message.reply_text(text[i:i+4000])
        return

    if update.message.document:
        fname = update.message.document.file_name or ''
        if fname.endswith('.xlsx'):
            await update.message.reply_text('Завантажую файл...')
            file = await ctx.bot.get_file(update.message.document.file_id)
            await file.download_to_drive(DATA_FILE)
            events = load_events()
            await update.message.reply_text(
                f'✅ Файл збережено! Знайдено {len(events)} подій.\n\n'
                f'Команди:\n'
                f'/today — найближчі 7 днів\n'
                f'/check — надіслати нагадування в групу\n'
                f'/list — всі події'
            )
            return

    if msg_text and chat_id != GROUP_CHAT_ID:
        await update.message.reply_text('Виконую...')
        result = subprocess.run(
            ['claude', '-p', msg_text, '--allowedTools', 'Bash,Read,Write,Edit', '--output-format', 'text'],
            capture_output=True, text=True, timeout=120,
            env={**os.environ, 'ANTHROPIC_API_KEY': ANTHROPIC_KEY}
        )
        reply = result.stdout or result.stderr or 'Немає відповіді'
        for i in range(0, len(reply), 4000):
            await update.message.reply_text(reply[i:i+4000])

app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()
app.add_handler(MessageHandler(filters.ALL, handle))
app.run_polling()
